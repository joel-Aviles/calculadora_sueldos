from funcs.converter import *
from openpyxl import Workbook
from datetime import datetime as dt


# Main
if __name__ == '__main__':

    # 1) Leer archivo
    df = converter(input("Nombre del archivo: "))

    # 2) Leer percepciones
    percep = converter("percepciones.xlsx")
    ## 2.1 Separar percepciones del archivo
    perord = df[(
        (df["tipoconcepto"] == "Percepción")
            &
        (df["conceptosiapsep"].isin(percep["clave"]))
    )]
    perext = df[(
        (df["tipoconcepto"] == "Percepción")
            &
        (~df["conceptosiapsep"].isin(percep["clave"]))
    )]

    # 3) Leer deducciones
    ## 3.1 Separar deducciones del archivo
    deducs = converter("deducciones.xlsx")
    deducs["concepto"] = deducs["concepto"].astype("string")
    gended = df[(
        (df["tipoconcepto"] == "Deducción")
            &
        (df["conceptosiapsep"].isin(deducs["concepto"]))
    )]
    outded = df[(
        (df["tipoconcepto"] == "Deducción")
            &
        (~df["conceptosiapsep"].isin(deducs["concepto"]))
    )]
    
    # 4) Datos para desplegar en el archivo
    ## 4.1 Sakr la primera y última quincena
    fstqnaproc = df["qnaproc"].min()
    lstqnaproc = df["qnaproc"].max()
    ## 4.2 Sakr RFC de la persona
    rfc = df["rfc"].unique()[0]
    ## 4.3 Sakr nombre de la persona
    ### 4.3.1 Leer archivo con las personas de la secretaría
    personal = converter("Personal.xlsx")
    ### 4.3.2 Asignar nombre según el RFC
    name = f"{personal.loc[personal["rfc"] == rfc, "nombre"].iloc[0].strip()} {personal.loc[personal["rfc"] == rfc, "paterno"].iloc[0].strip()} {personal.loc[personal["rfc"] == rfc, "materno"].iloc[0].strip()}"

    # 5) Pasar información a Excel
    ## 5.1 Inicializar escritores para Excel
    wb = Workbook() # Contenedor para la información del archivo
    ws = wb.active # Tomar hoja activa para escribir en ella
    ws.title = "Prueba" # Asignar nombre a nuestra hoja
    ## 5.2 Escribir información de la persona en el archivo
    ws["B2"] = "Datos del Empleado"
    ws["B3"] = "Nombre"
    ws["B4"] = "RFC"
    ws["B5"] = "Qna Devengada"
    ws["C3"] = name
    ws["C4"] = rfc
    ws["C5"] = lstqnaproc
    ## 5.3 Percepciones Ordinarias
    ### 5.3.1 Escribir en celdas para ir dando forma
    ws["B7"] = f"Percepciones Ordinarias {lstqnaproc}"
    ws["B8"] = "Consecutivo"
    ws["C8"] = "Tipo Nómina"
    ws["D8"] = "Tipo"
    ws["E8"] = "Clave"
    ws["F8"] = "Concepto"
    ws["G8"] = "Importe"
    ### 5.3.2 Inicializar variables contables y lista
    consec = 1 # Indicará el consecutivo del concepto
    xindex = 9 # Indica en qué fila vamos
    tempAr = [] # Guardará concepto, descripción y suma
    ### 5.3.3 Guardar percepciones ordinarias y sus cálculos
    for concepto in perord["conceptosiapsep"].unique():
        tempCo = {
            "concepto" : concepto,
            "descrip" : percep.loc[percep["clave"] == concepto, "descripcion"].iloc[0],
            "suma" : sum(
                perord[
                    (perord["conceptosiapsep"] == concepto)
                        &
                    (perord["qnaproc"] == lstqnaproc)
                ]["importe"]
            )
        }
        
        tempAr.append(tempCo)
    ### 5.3.4 Escribir percepciones ordinarias en el archivo
    for idx in range(len(tempAr)):
        if tempAr[idx]["suma"] != 0:
            ws[f"B{xindex}"] = consec
            ws[f"C{xindex}"] = "Ordinario"
            ws[f"D{xindex}"] = "Percepción"
            ws[f"E{xindex}"] = tempAr[idx]["concepto"]
            ws[f"F{xindex}"] = tempAr[idx]["descrip"]
            ws[f"G{xindex}"] = "{0:.2f}".format(tempAr[idx]["suma"])

            consec += 1
            xindex += 1
    ### 5.3.5 Agregar total de las percepciones al archivo
    ws[f"B{xindex}"] = "Total Percepciones Ordinarias"
    ws[f"G{xindex}"] = sum(item["suma"] for item in tempAr)
    ## 5.4 Reestablecer variables contables y lista
    consec = 1
    xindex += 2
    tempAr = []
    ## 5.5 Deducciones Ordinarias
    ### 5.5.1 Escribir en celdas para ir dando forma
    ws[f"B{xindex}"] = f"Deducciones Ordinarias {lstqnaproc}"
    
    xindex += 1
    
    ws[f"B{xindex}"] = "Consecutivo"
    ws[f"C{xindex}"] = "Tipo Nómina"
    ws[f"D{xindex}"] = "Tipo"
    ws[f"E{xindex}"] = "Clave"
    ws[f"F{xindex}"] = "Concepto"
    ws[f"G{xindex}"] = "Importe"
    
    xindex += 1
    ### 5.5.2 Guardar deducciones ordinarias y sus cálculos
    for concepto in gended["conceptosiapsep"].unique():
        tempCo = {
            "concepto" : concepto,
            "descrip" : deducs.loc[deducs["concepto"] == concepto, "descripcion"].iloc[0],
            "suma" : sum(
                gended[
                    (gended["conceptosiapsep"] == concepto)
                        &
                    (gended["qnaproc"] == lstqnaproc)
                ]["importe"]
            )
        }

        tempAr.append(tempCo)
    ### 5.5.3 Escribir deducciones ordinarias en el archivo
    for idx in range(len(tempAr)):
        if tempAr[idx]["suma"] != 0:
            ws[f"B{xindex}"] = consec
            ws[f"C{xindex}"] = "Ordinario"
            ws[f"D{xindex}"] = "Deducción"
            ws[f"E{xindex}"] = tempAr[idx]["concepto"]
            ws[f"F{xindex}"] = tempAr[idx]["descrip"]
            ws[f"G{xindex}"] = "{0:.2f}".format(tempAr[idx]["suma"])

            consec += 1
            xindex += 1
    ### 5.5.4 Agregar total de las deducciones al archivo
    ws[f"B{xindex}"] = "Total Deducciones Ordinarias"
    ws[f"G{xindex}"] = sum(item["suma"] for item in tempAr)
    ## 5.6 Reestablecer variables contables y lista
    consec = 1
    xindex += 2
    tempAr = []
    ## 5.7 Percepciones extraordinarias anuales
    ### 5.7.1 Escribir en celdas para ir dando forma
    ws[f"B{xindex}"] = f"Percepciones extraordinarias anuales"
    
    xindex += 1
    
    ws[f"B{xindex}"] = "Consecutivo"
    ws[f"C{xindex}"] = "Tipo Nómina"
    ws[f"D{xindex}"] = "Tipo"
    ws[f"E{xindex}"] = "Clave"
    ws[f"F{xindex}"] = "Concepto"
    ws[f"G{xindex}"] = "Importe"
    
    xindex += 1
    ### 5.7.2 Guardar percepciones extraordinarias anuales y sus cálculos
    #### 5.7.2.1 Obtener percepciones extra que no se toman en cuenta
    nocont = converter("PercepExtra_NoContarPensiones.xlsx")
    #### 5.7.2.2 Quedarse sólo con las que se toman en cuenta

    for concepto in perext["conceptosiapsep"].unique():
        if concepto not in nocont[nocont["cuenta"].str.lower() == "no"]["concepto"].to_list():
            tempCo = {
                "concepto" : concepto,
                "descrip" : perext.loc[perext["conceptosiapsep"] == concepto, "descripciondeconcepto"].iloc[0],
                "suma" : sum(
                    perext[
                        (perext["conceptosiapsep"] == concepto)
                            &
                        (perext["qnapago"] >= fstqnaproc)
                    ]["importe"]
                )
            }

            tempAr.append(tempCo)
    ### 5.7.3 Escribir percepciones extraordinarias anuales en el archivo
    for idx in range(len(tempAr)):
        if tempAr[idx]["suma"] != 0:
            ws[f"B{xindex}"] = consec
            ws[f"C{xindex}"] = "Extraordinario"
            ws[f"D{xindex}"] = "Percepción"
            ws[f"E{xindex}"] = tempAr[idx]["concepto"]
            ws[f"F{xindex}"] = tempAr[idx]["descrip"]
            ws[f"G{xindex}"] = "{0:.2f}".format(tempAr[idx]["suma"])

            consec += 1
            xindex += 1
    ### 5.7.4 Agregar total de las percepciones extraordinarias anuales al archivo
    ws[f"B{xindex}"] = "Total Percepciones Extraordinarias"
    ws[f"G{xindex}"] = sum(item["suma"] for item in tempAr)

    # 6) Guardar archivo
    ## 6.1 Tomar carpeta User de la cuenta activa
    userpath = os.path.expanduser(os.getenv('USERPROFILE'))
    ## 6.2 Wachar si existe el directorio al que se enviará
    dirpath = check_valid_dir(f"{userpath}/OneDrive - Secretaría de Educación de Guanajuato/tmp/Pensiones/{rfc}")
    ## 6.3 Chekr cuántos archivos de calculadora lleva
    counters = len([file for file in os.listdir(dirpath) if f"{rfc}-" in file and ".xlsx" in file])
    ## 6.4 Establecer nombre
    filename = f"{dirpath}/{rfc}-{dt.now().strftime("%d%m%Y")}_{counters + 1}.xlsx"
    ## 6.5 Guardarxd
    try:
        wb.save(filename=filename)
        print(f"Archivo guardado exitosamente con el nombre '{filename.split("/")[-1]}' uwu")
    except Exception as ex:
        print(f"Error = {ex}")
    finally:
        print("Programa Finalizado")

    