from openpyxl import Workbook
from datetime import datetime as dt
from unidecode import unidecode
import pandas as pd
import os
import sys
import re

# FUNCIÓN PARA BR SI EXISTE UN DIRECTORIO
def check_valid_dir(dirpath):
    if not os.path.exists(dirpath):
        os.makedirs(dirpath)
        return dirpath
    else:
        return dirpath

# FUNCIÓN PARA TENER UN PATH VÁLIDO PARA PANDAS
def valid_path(path):
    return re.sub("\s*", "", re.sub('[&%"]', '', path.replace("\\", "/")).replace("'", ""), count=1)

# FUNCIÓN PARA TENER NOMBRES DE COLUMNAS FÁCILES DE MANEJAR
def validname(name):
    return unidecode(''.join(filter(str.isalpha, name)).lower())

# FUNCIÓN PRINCIPAL -> CONVERTIR ARCHIVOS A PANDA'S DATAFRAME
def converter(filename, sheetname="Hoja1"):

    dataframe = pd.DataFrame(pd.read_excel(valid_path(filename)))
    dataframe.dropna(how="all", inplace=True)
    dataframe.columns = [validname(column) for column in dataframe.columns]

    return dataframe

# Definir constantes
LEY_CONCEPTS = ["01", "58", "59", "5G", "70", "74", "77", "IN", "P2", "R9", "RV", "SS"]
SUELDO_CONCEPTS = ["07", "7A", "7B", "7C", "7D", "7E"]

def get_resource_path(relative_path):
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def load_excel_file(filename):
    """Carga un archivo Excel desde la ruta correcta, usando get_resource_path."""
    file_path = get_resource_path(f'excel_files/{filename}')
    df_file = converter(file_path)
    return df_file

def calculate_amounts(percep_ord, percep_extra, deductions, ley_deductions, sueldo, formula_type):
    formulas = {
        '1': {'formula_name': 'Liquido', 'formula': lambda: (percep_ord + percep_extra) - ley_deductions},
        '2': {'formula_name': 'Neto', 'formula': lambda: (percep_ord + percep_extra) - deductions},
        '3': {'formula_name': 'Solo percepciones', 'formula': lambda: percep_ord + percep_extra},
        '4': {'formula_name': 'Percepciones ordinarias - deducciones de ley', 'formula': lambda: percep_ord - ley_deductions},
        '5': {'formula_name': 'Percepciones extraordinarias - deducciones de ley', 'formula': lambda: percep_extra - ley_deductions},
        '6': {'formula_name': 'Solo percepciones ordinarias', 'formula': lambda: percep_ord},
        '7': {'formula_name': 'Solo percepciones extraordinarias', 'formula': lambda: percep_extra},
        '8': {'formula_name': 'Sueldo (percepciones 07)', 'formula': lambda: sueldo},
    }

    selected_formula = formulas.get(formula_type, formulas['1'])
    return {
        'formula_name': selected_formula['formula_name'],
        'amount': selected_formula['formula']()
    }

def get_personal_data(df, personal):
    rfc = df["rfc"].unique()[0]
    person = personal.loc[personal["rfc"] == rfc].iloc[0]
    name = f"{person['nombre'].strip()} {person['paterno'].strip()} {person['materno'].strip()}"
    return rfc, name

def write_excel(ws, section, data, start_row, total_label):
    row = start_row
    ws[f"B{row}"] = section
    row += 1
    ws[f"B{row}"] = "Consecutivo"
    ws[f"C{row}"] = "Tipo Nómina"
    ws[f"D{row}"] = "Tipo"
    ws[f"E{row}"] = "Clave"
    ws[f"F{row}"] = "Concepto"
    ws[f"G{row}"] = "Importe"
    row += 1

    total_sum = 0
    for consec, item in enumerate(data, start=1):
        if item["suma"] != 0:
            ws[f"B{row}"] = consec
            ws[f"C{row}"] = "Ordinario" if "Ordinaria" in section else "Extraordinario"
            ws[f"D{row}"] = item["tipo"]
            ws[f"E{row}"] = item["concepto"]
            ws[f"F{row}"] = item["descrip"]
            ws[f"G{row}"] = "{0:.2f}".format(item["suma"])
            total_sum += item["suma"]
            row += 1

    ws[f"B{row}"] = total_label
    ws[f"G{row}"] = total_sum
    return row + 2, total_sum

def prompt_user_input():
    process_type = input("¿Qué proceso deseas llevar a cabo?\n 1) Pensión alimenticia\n 2) Juicios mercantiles\nIngrese su elección (1 ó 2): ")
    
    if process_type == '1':
        discount_percent = int(input("¿Qué porcentaje se le va a descontar al trabajador? ")) * 0.01
        money_formula = input("¿Qué fórmula se usará?\n 1) Líquido\n 2) Neto\n 3) Solo percepciones\n 4) Percepciones ordinarias - deducciones de ley\n 5) Percepciones extraordinarias - deducciones de ley\n 6) Solo ordinarias\n 7) Solo extraordinarias\n 8) Percepciones 07\nIngrese el número de su elección (default: 1): ")
        payment_period = int(input("¿Cuántas quincenas se le va a cobrar? "))
    else:
        discount_percent = 1
        money_formula = '1'
        payment_period = 1

    file_name = input("Nombre del archivo: ")
    
    return process_type, discount_percent, money_formula, payment_period, file_name

def main():
    # 1) Leer archivo
    process_type, discount_percent, money_formula, payment_period, file_name = prompt_user_input()
    df = converter(file_name)

    # 2) Leer percepciones y deducciones
    percep = load_excel_file("percepciones.xlsx")
    perord = df[(df["tipoconcepto"] == "Percepción") & (df["conceptosiapsep"].isin(percep["clave"]))]
    perext = df[(df["tipoconcepto"] == "Percepción") & (~df["conceptosiapsep"].isin(percep["clave"]))]

    deducs = load_excel_file("deducciones.xlsx")
    deducs["concepto"] = deducs["concepto"].astype(str)
    gended = df[(df["tipoconcepto"] == "Deducción") & (df["conceptosiapsep"].isin(deducs["concepto"]))]
    outded = df[(df["tipoconcepto"] == "Deducción") & (~df["conceptosiapsep"].isin(deducs["concepto"]))]

    # 4) Obtener datos del empleado
    fstqnaproc = df["qnaproc"].min()
    lstqnaproc = df["qnaproc"].max()
    personal = load_excel_file("Personal.xlsx")
    rfc, name = get_personal_data(df, personal)

    # 5) Crear y escribir en archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Prueba"

    ws["B2"] = "Datos del Empleado"
    ws["B3"] = "Nombre"
    ws["B4"] = "RFC"
    ws["B5"] = "Qna Devengada"
    ws["C3"] = name
    ws["C4"] = rfc
    ws["C5"] = lstqnaproc

    # Percepciones Ordinarias
    data = [
        {
            "concepto": concepto,
            "descrip": percep.loc[percep["clave"] == concepto, "descripcion"].iloc[0],
            "suma": perord[(perord["conceptosiapsep"] == concepto) & (perord["qnaproc"] == lstqnaproc)]["importe"].sum(),
            "tipo": "Percepción"
        }
        for concepto in perord["conceptosiapsep"].unique()
    ]
    xindex, total_percep_ord = write_excel(ws, f"Percepciones Ordinarias {lstqnaproc}", data, 7, "Total Percepciones Ordinarias")

    # Deducciones Ordinarias
    data = [
        {
            "concepto": concepto,
            "descrip": deducs.loc[deducs["concepto"] == concepto, "descripcion"].iloc[0],
            "suma": gended[(gended["conceptosiapsep"] == concepto) & (gended["qnaproc"] == lstqnaproc)]["importe"].sum(),
            "tipo": "Deducción"
        }
        for concepto in gended["conceptosiapsep"].unique()
    ]
    xindex, total_deduc_ord = write_excel(ws, f"Deducciones Ordinarias {lstqnaproc}", data, xindex, "Total Deducciones Ordinarias")

    # Percepciones Extraordinarias
    nocont = load_excel_file("PercepExtra_NoContarPensiones.xlsx")
    data = [
        {
            "concepto": concepto,
            "descrip": perext.loc[perext["conceptosiapsep"] == concepto, "descripciondeconcepto"].iloc[0],
            "suma": perext[(perext["conceptosiapsep"] == concepto) & (perext["qnapago"] >= fstqnaproc)]["importe"].sum(),
            "tipo": "Percepción"
        }
        for concepto in perext["conceptosiapsep"].unique()
        if concepto not in nocont[nocont["cuenta"].str.lower() == "no"]["concepto"].to_list()
    ]
    xindex, total_percep_extra = write_excel(ws, "Percepciones extraordinarias anuales", data, xindex, "Total Percepciones Extraordinarias")

    # Deducciones de ley
    total_deduc_ley = sum(
        gended[(gended["conceptosiapsep"] == concepto) & (gended["qnaproc"] == lstqnaproc)]["importe"].sum()
        for concepto in gended["conceptosiapsep"].unique()
        if concepto in LEY_CONCEPTS
    )

    # Percepciones de sueldo
    total_sueldo = sum(
        perord[(perord["conceptosiapsep"] == concepto) & (perord["qnaproc"] == lstqnaproc)]["importe"].sum()
        for concepto in perord["conceptosiapsep"].unique()
        if concepto in SUELDO_CONCEPTS
    )

    # 6) Formato para pensiones alimenticias
    if process_type == '1':
        formula_result = calculate_amounts(total_percep_ord, total_percep_extra, total_deduc_ord, total_deduc_ley, total_sueldo, money_formula)
        mount_to_discount = formula_result["amount"] * discount_percent

        ws[f"B{xindex}"] = "Fórmula usada"
        ws[f"C{xindex}"] = formula_result["formula_name"]
        xindex += 1
        ws[f"B{xindex}"] = "Descuento del " + str(discount_percent * 100) + "%"
        ws[f"C{xindex}"] = mount_to_discount
        xindex += 2

        # Pagos retroactivos
        ws[f"B{xindex}"] = "Periodo"
        ws[f"C{xindex}"] = str(payment_period) + " Quincenas"
        xindex += 1
        ws[f"B{xindex}"] = "Consecutivo"
        ws[f"C{xindex}"] = "Número de quincena"
        ws[f"D{xindex}"] = "Monto"
        xindex += 1

        mount_per_period = mount_to_discount / payment_period
        for i in range(payment_period):
            lstqna_str = str(lstqnaproc)

            year = lstqna_str[:4]
            num_qna = lstqna_str[4:]

            year = int(year)
            num_qna = int(num_qna)

            if (num_qna >= 24):
                year += 1
                num_qna = 0
            
            num_qna += 1

            lstqnaproc = f"{year}{num_qna:02}"

            ws[f"B{xindex}"] = i + 1
            ws[f"C{xindex}"] = lstqnaproc
            ws[f"D{xindex}"] = "{0:.2f}".format(mount_per_period)
            xindex += 1

    # 7) Guardar archivo Excel
    userpath = os.path.expanduser(os.getenv('USERPROFILE'))
    dirpath = check_valid_dir(f"{userpath}/OneDrive - Secretaría de Educación de Guanajuato/tmp/Pensiones/{rfc}")
    counters = len([file for file in os.listdir(dirpath) if f"{rfc}-" in file and ".xlsx" in file])
    filename = f"{dirpath}/{rfc}-{dt.now().strftime('%d%m%Y')}_{counters + 1}.xlsx"

    try:
        wb.save(filename=filename)
        print(f"Archivo guardado exitosamente con el nombre '{filename.split('/')[-1]}'")
    except Exception as ex:
        print(f"Error = {ex}")    

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"Se ha producido un error: {e}")
    finally:
        os.system("pause")
