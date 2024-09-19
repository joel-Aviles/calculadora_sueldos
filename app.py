from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import FileResponse
from funcs.converter import converter, check_valid_dir
from openpyxl import Workbook
from datetime import datetime as dt
import pandas as pd
import os

# Crear la instancia de FastAPI
app = FastAPI()

# Definir constantes
LEY_CONCEPTS = ["01", "58", "59", "5G", "70", "74", "77", "IN", "P2", "R9", "RV", "SS"]
SUELDO_CONCEPTS = ["07", "7A", "7B", "7C", "7D", "7E"]

def load_excel_file(file: UploadFile):
    df_file = converter(file.file)
    return df_file

def get_personal_data(df, personal):
    rfc = df["rfc"].unique()[0]
    person = personal.loc[personal["rfc"] == rfc].iloc[0]
    name = f"{person['nombre'].strip()} {person['paterno'].strip()} {person['materno'].strip()}"
    return rfc, name

def write_excel(ws, section, data, start_row, total_label):
    row = start_row
    ws[f"B{row}"] = section
    row += 1
    ws[f"B{row}"] = "Consecutivo"
    ws[f"C{row}"] = "Tipo Nómina"
    ws[f"D{row}"] = "Tipo"
    ws[f"E{row}"] = "Clave"
    ws[f"F{row}"] = "Concepto"
    ws[f"G{row}"] = "Importe"
    row += 1

    total_sum = 0
    for consec, item in enumerate(data, start=1):
        if item["suma"] != 0:
            ws[f"B{row}"] = consec
            ws[f"C{row}"] = "Ordinario" if "Ordinaria" in section else "Extraordinario"
            ws[f"D{row}"] = item["tipo"]
            ws[f"E{row}"] = item["concepto"]
            ws[f"F{row}"] = item["descrip"]
            ws[f"G{row}"] = "{0:.2f}".format(item["suma"])
            total_sum += item["suma"]
            row += 1

    ws[f"B{row}"] = total_label
    ws[f"G{row}"] = total_sum
    return row + 2, total_sum

def calculate_amounts(percep_ord, percep_extra, deductions, ley_deductions, sueldo, formula_type):
    formulas = {
        '1': {'formula_name': 'Liquido', 'formula': lambda: (percep_ord + percep_extra) - ley_deductions},
        '2': {'formula_name': 'Neto', 'formula': lambda: (percep_ord + percep_extra) - deductions},
        '3': {'formula_name': 'Solo percepciones', 'formula': lambda: percep_ord + percep_extra},
        '4': {'formula_name': 'Percepciones ordinarias - deducciones de ley', 'formula': lambda: percep_ord - ley_deductions},
        '5': {'formula_name': 'Percepciones extraordinarias - deducciones de ley', 'formula': lambda: percep_extra - ley_deductions},
        '6': {'formula_name': 'Solo percepciones ordinarias', 'formula': lambda: percep_ord},
        '7': {'formula_name': 'Solo percepciones extraordinarias', 'formula': lambda: percep_extra},
        '8': {'formula_name': 'Sueldo (percepciones 07)', 'formula': lambda: sueldo},
    }

    selected_formula = formulas.get(formula_type, formulas['1'])
    return {
        'formula_name': selected_formula['formula_name'],
        'amount': selected_formula['formula']()
    }

# Endpoint para procesar archivo
@app.post("/procesar/")
async def procesar_archivo(
    process_type: str = Form(...),
    discount_percent: float = Form(...),
    money_formula: str = Form(...),
    payment_period: int = Form(...),
    file: UploadFile = File(...)
):
    # 1) Leer archivo subido
    df = converter(file.file)

    # 2) Leer percepciones y deducciones
    percep = load_excel_file(await File("percepciones.xlsx"))
    perord = df[(df["tipoconcepto"] == "Percepción") & (df["conceptosiapsep"].isin(percep["clave"]))]
    perext = df[(df["tipoconcepto"] == "Percepción") & (~df["conceptosiapsep"].isin(percep["clave"]))]

    deducs = load_excel_file(await File("deducciones.xlsx"))
    deducs["concepto"] = deducs["concepto"].astype(str)
    gended = df[(df["tipoconcepto"] == "Deducción") & (df["conceptosiapsep"].isin(deducs["concepto"]))]

    # 4) Obtener datos del empleado
    fstqnaproc = df["qnaproc"].min()
    lstqnaproc = df["qnaproc"].max()
    personal = load_excel_file(await File("Personal.xlsx"))
    rfc, name = get_personal_data(df, personal)

    # 5) Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Prueba"

    ws["B2"] = "Datos del Empleado"
    ws["C3"] = name
    ws["C4"] = rfc

    # Percepciones Ordinarias
    data = [
        {
            "concepto": concepto,
            "descrip": percep.loc[percep["clave"] == concepto, "descripcion"].iloc[0],
            "suma": perord[(perord["conceptosiapsep"] == concepto) & (perord["qnaproc"] == lstqnaproc)]["importe"].sum(),
            "tipo": "Percepción"
        }
        for concepto in perord["conceptosiapsep"].unique()
    ]
    xindex, total_percep_ord = write_excel(ws, f"Percepciones Ordinarias {lstqnaproc}", data, 7, "Total Percepciones Ordinarias")

    # Deducciones Ordinarias
    data = [
        {
            "concepto": concepto,
            "descrip": deducs.loc[deducs["concepto"] == concepto, "descripcion"].iloc[0],
            "suma": gended[(gended["conceptosiapsep"] == concepto) & (gended["qnaproc"] == lstqnaproc)]["importe"].sum(),
            "tipo": "Deducción"
        }
        for concepto in gended["conceptosiapsep"].unique()
    ]
    xindex, total_deduc_ord = write_excel(ws, f"Deducciones Ordinarias {lstqnaproc}", data, xindex, "Total Deducciones Ordinarias")

    # Percepciones Extraordinarias
    nocont = load_excel_file("PercepExtra_NoContarPensiones.xlsx")
    data = [
        {
            "concepto": concepto,
            "descrip": perext.loc[perext["conceptosiapsep"] == concepto, "descripciondeconcepto"].iloc[0],
            "suma": perext[(perext["conceptosiapsep"] == concepto) & (perext["qnapago"] >= fstqnaproc)]["importe"].sum(),
            "tipo": "Percepción"
        }
        for concepto in perext["conceptosiapsep"].unique()
        if concepto not in nocont[nocont["cuenta"].str.lower() == "no"]["concepto"].to_list()
    ]
    xindex, total_percep_extra = write_excel(ws, "Percepciones extraordinarias anuales", data, xindex, "Total Percepciones Extraordinarias")

    # Deducciones de ley
    total_deduc_ley = sum(
        gended[(gended["conceptosiapsep"] == concepto) & (gended["qnaproc"] == lstqnaproc)]["importe"].sum()
        for concepto in gended["conceptosiapsep"].unique()
        if concepto in LEY_CONCEPTS
    )

    # Percepciones de sueldo
    total_sueldo = sum(
        perord[(perord["conceptosiapsep"] == concepto) & (perord["qnaproc"] == lstqnaproc)]["importe"].sum()
        for concepto in perord["conceptosiapsep"].unique()
        if concepto in SUELDO_CONCEPTS
    )

    # 6) Formato para pensiones alimenticias
    if process_type == '1':
        formula_result = calculate_amounts(total_percep_ord, total_percep_extra, total_deduc_ord, total_deduc_ley, total_sueldo, money_formula)
        mount_to_discount = formula_result["amount"] * discount_percent

        ws[f"B{xindex}"] = "Fórmula usada"
        ws[f"C{xindex}"] = formula_result["formula_name"]
        xindex += 1
        ws[f"B{xindex}"] = "Descuento del " + str(discount_percent * 100) + "%"
        ws[f"C{xindex}"] = mount_to_discount
        xindex += 2

        # Pagos retroactivos
        ws[f"B{xindex}"] = "Periodo"
        ws[f"C{xindex}"] = str(payment_period) + " Quincenas"
        xindex += 1
        ws[f"B{xindex}"] = "Consecutivo"
        ws[f"C{xindex}"] = "Número de quincena"
        ws[f"D{xindex}"] = "Monto"
        xindex += 1

        mount_per_period = mount_to_discount / payment_period
        for i in range(payment_period):
            lstqna_str = str(lstqnaproc)

            year = lstqna_str[:4]
            num_qna = lstqna_str[4:]

            year = int(year)
            num_qna = int(num_qna)

            if (num_qna >= 24):
                year += 1
                num_qna = 0
            
            num_qna += 1

            lstqnaproc = f"{year}{num_qna:02}"

            ws[f"B{xindex}"] = i + 1
            ws[f"C{xindex}"] = lstqnaproc
            ws[f"D{xindex}"] = "{0:.2f}".format(mount_per_period)
            xindex += 1

    # 7) Guardar archivo en un directorio temporal
    userpath = os.path.expanduser(os.getenv('USERPROFILE') or "~")
    dirpath = check_valid_dir(f"{userpath}/tmp/Pensiones/{rfc}")
    if not os.path.exists(dirpath):
        os.makedirs(dirpath)
    filename = f"{dirpath}/{rfc}-{dt.now().strftime('%d%m%Y')}.xlsx"
    wb.save(filename)

    return {"message": "Archivo procesado exitosamente", "filename": filename}

# Endpoint para descargar el archivo generado
@app.get("/descargar/")
async def descargar_archivo(filename: str):
    # Verificar si el archivo existe
    if not os.path.exists(filename):
        return {"error": "El archivo no existe."}
    
    # Retornar el archivo como respuesta
    return FileResponse(path=filename, filename=os.path.basename(filename), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.get("/health")
def read_root():
    return {"health": "ok"}

# Ejecutar la aplicación
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
